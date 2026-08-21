"""Import reviewed T/CHNRISC energy-quota tables into a local SQLite database.

The PDF and generated database are intentionally ignored by Git. This importer
contains parsing logic and provenance checks, not a bundled copy of the source.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from dataclasses import replace
from pathlib import Path

try:
    import pdfplumber
except ImportError as exc:  # pragma: no cover - operator environment guard
    raise SystemExit("pdfplumber is required to import the standard PDF") from exc

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional workbook import guard
    openpyxl = None
    get_column_letter = None

from a1_factor_engine import (
    EnergyConversionRecord,
    EnergyQuotaModifierRule,
    EnergyQuotaRecord,
    EnterpriseEnergyProfileRecord,
    EnterpriseProcessEmissionRecord,
    ParameterSourceType,
    ScopedProcessParameterRecord,
    create_energy_database,
)
from a1_factor_engine.material_registry import DEFAULT_MATERIAL_REGISTRY

STANDARD_CODE = "T/CHNRISC 0008-2025"
PUBLISHER = "河南省耐火材料行业协会"
WATERMARK_PARTS = frozenset("平台信息标准团体全国")
ENTERPRISE_SHEET_NAME = "能碳转换碳排放核算--89个品种"
ENTERPRISE_PROVIDER = "企业能耗核算数据汇编（用户提供）"
REVIEW_MARKERS = re.compile(
    r"(?:需要核实|核实|没有企业数据|无企业数据|征求意见|参考|是否|合并|"
    r"类比|\?|？|待确认|同上)"
)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_cell(value: object) -> str:
    lines = [line.strip() for line in str(value or "").replace("\r", "\n").split("\n")]
    return " ".join(line for line in lines if line and line not in WATERMARK_PARTS).strip()


def numeric(value: object) -> float | None:
    observed = clean_cell(value).replace(" ", "")
    return float(observed) if re.fullmatch(r"\d+(?:\.\d+)?", observed) else None


def _slug(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    return normalized or hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]


def canonical_product_key(product_name: str) -> str:
    return {
        "电熔莫来石": "electrofused mullite",
        "烧结莫来石": "sintered mullite",
        "烧结矾土基莫来石": "sintered bauxite-based mullite",
        "电熔锆莫来石": "electrofused zircon mullite",
        "电熔镁铝尖晶石": "electrofused spinel",
        "烧结镁铝尖晶石": "sintered spinel",
    }.get(product_name, product_name)


def _workbook_number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _workbook_text(value: object) -> str:
    return " ".join(str(value or "").replace("\r", "\n").split())


def extract_enterprise_profiles(
    workbook_path: Path,
    source_sha: str,
) -> list[EnterpriseEnergyProfileRecord]:
    """Extract current-data rows and preserve level-specific allocation provenance."""

    if openpyxl is None or get_column_letter is None:
        raise SystemExit("openpyxl is required for --enterprise-energy-workbook")
    formulas_book = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    values_book = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        if ENTERPRISE_SHEET_NAME not in formulas_book.sheetnames:
            raise ValueError(f"enterprise workbook is missing sheet {ENTERPRISE_SHEET_NAME}")
        formula_sheet = formulas_book[ENTERPRISE_SHEET_NAME]
        value_sheet = values_book[ENTERPRISE_SHEET_NAME]

        def is_primary(row_number: int) -> bool:
            values = tuple(
                _workbook_number(value_sheet.cell(row_number, column).value)
                for column in range(4, 8)
            )
            has_label = bool(
                _workbook_text(value_sheet.cell(row_number, 2).value)
                or _workbook_text(value_sheet.cell(row_number, 3).value)
            )
            return bool(
                has_label
                and all(value is not None for value in values)
                and abs(values[3] - 0.5306) <= 1e-9
            )

        primary_rows = [
            row_number for row_number in range(5, value_sheet.max_row + 1)
            if is_primary(row_number)
        ]
        output: list[EnterpriseEnergyProfileRecord] = []
        product_group = ""
        sequence_id = ""
        for position, row_number in enumerate(primary_rows):
            raw_sequence = _workbook_number(value_sheet.cell(row_number, 1).value)
            if raw_sequence is not None:
                sequence_id = f"{raw_sequence:g}"
            observed_group = _workbook_text(value_sheet.cell(row_number, 2).value)
            if observed_group:
                product_group = observed_group
            variant = _workbook_text(value_sheet.cell(row_number, 3).value)
            product_name = variant or product_group
            if not sequence_id or not product_name:
                raise ValueError(f"enterprise profile row {row_number} has incomplete identity")
            combined_name = " ".join(dict.fromkeys(filter(None, (product_group, variant))))
            mapped_product = canonical_product_key(product_name)
            canonical_product = (
                mapped_product if mapped_product != product_name else combined_name
            )
            resolved = DEFAULT_MATERIAL_REGISTRY.resolve(combined_name or product_name)
            identity = resolved.identity
            head_material = identity.head_material or product_name
            process = identity.manufacturing_route[0] if identity.manufacturing_route else ""

            next_row = primary_rows[position + 1] if position + 1 < len(primary_rows) else value_sheet.max_row + 1
            context: list[str] = []
            for context_row in range(row_number + 1, next_row):
                for column in (2, 3, 22, 23):
                    text = _workbook_text(value_sheet.cell(context_row, column).value)
                    if not text or text in {"0005能耗团标", "0006碳排放团标", "无，新增"}:
                        continue
                    if text.startswith("能耗标准修订后"):
                        continue
                    context.append(f"{get_column_letter(column)}{context_row}: {text}")
            primary_note = _workbook_text(value_sheet.cell(row_number, 22).value)
            note_parts = tuple(filter(None, (primary_note, *context)))
            quality_note = " | ".join(note_parts) or "workbook allocation; no additional row note"
            needs_review = bool(REVIEW_MARKERS.search(quality_note))

            for level, (energy_column, share_column, formula_column) in enumerate(
                ((4, 10, 13), (5, 11, 14), (6, 12, 15)),
                1,
            ):
                total_energy = _workbook_number(value_sheet.cell(row_number, energy_column).value)
                electricity_share = _workbook_number(value_sheet.cell(row_number, share_column).value)
                if total_energy is None or electricity_share is None:
                    raise ValueError(
                        f"enterprise profile row {row_number} level {level} lacks energy/share"
                    )
                formula = _workbook_text(formula_sheet.cell(row_number, formula_column).value)
                remainder_share = 1.0 - electricity_share
                if abs(remainder_share) <= 1e-12:
                    remainder_share = 0.0
                    remainder_carrier = "none"
                    allocation_status = "ALL_ELECTRIC"
                elif (
                    re.search(rf"\bI{row_number}\b", formula, re.IGNORECASE)
                    and _workbook_number(value_sheet.cell(row_number, 9).value) is not None
                ):
                    remainder_carrier = "standard_coal"
                    allocation_status = "ELECTRICITY_STANDARD_COAL"
                elif (
                    re.search(rf"\bH{row_number}\b", formula, re.IGNORECASE)
                    and _workbook_number(value_sheet.cell(row_number, 8).value) is not None
                ):
                    remainder_carrier = "natural_gas"
                    allocation_status = "ELECTRICITY_NATURAL_GAS"
                else:
                    remainder_carrier = "unresolved"
                    allocation_status = "UNRESOLVED_REMAINDER"
                supported_carrier = remainder_carrier in {"none", "natural_gas"}
                runtime_eligible = supported_carrier and not needs_review
                if needs_review:
                    allocation_status += "_NEEDS_REVIEW"
                energy_cell = f"{get_column_letter(energy_column)}{row_number}"
                share_cell = f"{get_column_letter(share_column)}{row_number}"
                formula_cell = f"{get_column_letter(formula_column)}{row_number}"
                output.append(EnterpriseEnergyProfileRecord(
                    profile_id=f"enterprise-energy-89:r{row_number:03d}:l{level}",
                    sequence_id=sequence_id,
                    product_name=product_name,
                    product_group=product_group,
                    head_material=head_material,
                    production_process=process,
                    product_form=identity.product_form or "",
                    quota_level=level,
                    total_energy_kgce_per_t=total_energy,
                    electricity_share=electricity_share,
                    remainder_carrier=remainder_carrier,
                    remainder_share=remainder_share,
                    source_type=ParameterSourceType.USER_CONFIRMED_ENGINEERING_DATA,
                    provider=ENTERPRISE_PROVIDER,
                    locator=(
                        f"workbook:{workbook_path.name}#'{ENTERPRISE_SHEET_NAME}'!"
                        f"{energy_cell}:{formula_cell}"
                    ),
                    worksheet_name=ENTERPRISE_SHEET_NAME,
                    worksheet_row=row_number,
                    energy_cell=energy_cell,
                    electricity_share_cell=share_cell,
                    formula_cell=formula_cell,
                    canonical_product=canonical_product,
                    citation=(
                        f"{workbook_path.name}，{ENTERPRISE_SHEET_NAME}，"
                        f"第{row_number}行，{level}级"
                    ),
                    quality_note=quality_note,
                    allocation_status=allocation_status,
                    source_sha256=source_sha,
                    metadata={
                        "combined_product_name": combined_name,
                        "formula": formula,
                        "primary_note": primary_note,
                        "context_json": json.dumps(context, ensure_ascii=False),
                        "electricity_factor_tco2_per_mwh": _workbook_text(
                            value_sheet.cell(row_number, 7).value
                        ),
                        "natural_gas_factor_tco2_per_tce": _workbook_text(
                            value_sheet.cell(row_number, 8).value
                        ),
                        "standard_coal_factor_tco2_per_tce": _workbook_text(
                            value_sheet.cell(row_number, 9).value
                        ),
                    },
                    runtime_eligible=runtime_eligible,
                ))
        if len(output) != 273:
            raise ValueError(f"expected 273 level profiles, observed {len(output)}")
        if len({item.sequence_id for item in output}) != 89:
            raise ValueError("enterprise workbook does not contain exactly 89 sequence IDs")
        key_counts: dict[tuple[str, int], int] = {}
        for item in output:
            key = (item.canonical_product.casefold(), item.quota_level)
            key_counts[key] = key_counts.get(key, 0) + 1
        return [
            replace(
                item,
                runtime_eligible=False,
                allocation_status=f"{item.allocation_status}_AMBIGUOUS_DUPLICATE",
                quality_note=(
                    f"{item.quality_note} | duplicate canonical product/level; "
                    "runtime selection requires a distinguishing attribute"
                ),
            )
            if key_counts[(item.canonical_product.casefold(), item.quota_level)] > 1
            else item
            for item in output
        ]
    finally:
        formulas_book.close()
        values_book.close()


def extract_enterprise_process_emissions(
    workbook_path: Path,
    source_sha: str,
    profiles: list[EnterpriseEnergyProfileRecord],
) -> list[EnterpriseProcessEmissionRecord]:
    """Extract level-specific non-energy process CO2 cells with exact workbook lineage."""

    if openpyxl is None or get_column_letter is None:
        raise SystemExit("openpyxl is required for --enterprise-energy-workbook")
    formulas_book = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    values_book = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    try:
        formula_sheet = formulas_book[ENTERPRISE_SHEET_NAME]
        value_sheet = values_book[ENTERPRISE_SHEET_NAME]
        output: list[EnterpriseProcessEmissionRecord] = []
        for profile in profiles:
            if not profile.production_process:
                continue
            emission_column = 15 + profile.quota_level
            raw_value = _workbook_number(
                value_sheet.cell(profile.worksheet_row, emission_column).value
            )
            if raw_value is None:
                continue
            emission_cell = f"{get_column_letter(emission_column)}{profile.worksheet_row}"
            formula = _workbook_text(
                formula_sheet.cell(profile.worksheet_row, emission_column).value
            )
            remark = _workbook_text(value_sheet.cell(profile.worksheet_row, 22).value)
            is_electrode = "电极" in remark or bool(re.search(r"44\s*/\s*12", formula))
            is_decomposition = "分解" in remark
            if is_electrode and is_decomposition:
                emission_name = "combined_electrode_and_decomposition_co2"
            elif is_electrode:
                emission_name = "direct_electrode_oxidation_co2"
            elif is_decomposition:
                emission_name = "decomposition_process_co2"
            else:
                emission_name = "additional_direct_process_co2"
            metadata = {
                "enterprise_energy_profile_id": profile.profile_id,
                "process_emission_kind": emission_name,
                "remark": remark,
                "raw_unit": "kgCO2e/t product",
                "stoichiometric_formula": formula,
            }
            output.append(EnterpriseProcessEmissionRecord(
                emission_id=(
                    f"enterprise-process-emission-89:r{profile.worksheet_row:03d}:"
                    f"l{profile.quota_level}:{emission_name}"
                ),
                sequence_id=profile.sequence_id,
                product_name=profile.product_name,
                canonical_product=profile.canonical_product,
                head_material=profile.head_material,
                production_process=profile.production_process,
                quota_level=profile.quota_level,
                emission_name=emission_name,
                value_kgco2e_per_t=raw_value,
                source_type=ParameterSourceType.USER_CONFIRMED_ENGINEERING_DATA,
                provider=ENTERPRISE_PROVIDER,
                locator=(
                    f"workbook:{workbook_path.name}#'{ENTERPRISE_SHEET_NAME}'!{emission_cell}"
                ),
                worksheet_name=ENTERPRISE_SHEET_NAME,
                worksheet_row=profile.worksheet_row,
                emission_cell=emission_cell,
                formula=formula,
                citation=(
                    f"{workbook_path.name}，{ENTERPRISE_SHEET_NAME}，"
                    f"{emission_cell}，{profile.quota_level}级"
                ),
                quality_note=(
                    f"{profile.quality_note} | additional process emission; "
                    f"remark: {remark or 'none'}"
                ),
                source_sha256=source_sha,
                metadata=metadata,
                runtime_eligible=profile.runtime_eligible,
            ))
        return output
    finally:
        formulas_book.close()
        values_book.close()


def extract_quotas(pdf: pdfplumber.PDF, source_sha: str) -> list[EnergyQuotaRecord]:
    specs = (
        (5, 0, "1"),
        (6, 0, "1"),
        (6, 1, "2"),
        (7, 0, "2"),
        (8, 1, "3"),
    )
    output: list[EnergyQuotaRecord] = []
    contexts: dict[str, list[str]] = {}
    row_numbers: dict[str, int] = {"1": 0, "2": 0, "3": 0}
    for page_index, table_index, table_number in specs:
        tables = pdf.pages[page_index].extract_tables()
        if table_index >= len(tables):
            raise ValueError(f"table {table_number} not found on physical page {page_index + 1}")
        context = contexts.setdefault(table_number, [])
        for raw in tables[table_index]:
            if len(raw) < 4:
                continue
            levels = tuple(numeric(value) for value in raw[-3:])
            labels = [clean_cell(value) for value in raw[:-3]]
            if any("注：" in label for label in labels) or levels == (None, None, None):
                continue
            if any(value is None for value in levels):
                continue
            while len(context) < len(labels):
                context.append("")
            for index, label in enumerate(labels):
                if label:
                    context[index] = label
                    for deeper in range(index + 1, len(context)):
                        context[deeper] = ""
            path = [value for value in context if value]
            if not path:
                raise ValueError(f"quota row has no product label on physical page {page_index + 1}")
            row_numbers[table_number] += 1
            row_number = row_numbers[table_number]
            product_name = path[-1]
            product_path = " / ".join(path)
            resolved = DEFAULT_MATERIAL_REGISTRY.resolve(product_path)
            identity = resolved.identity
            head_material = identity.head_material or product_name
            process = identity.manufacturing_route[0] if identity.manufacturing_route else ""
            locator = f"standard:{STANDARD_CODE.replace(' ', '-')}#table-{table_number}"
            for level, value in enumerate(levels, 1):
                output.append(EnergyQuotaRecord(
                    record_id=(
                        f"t-chnrisc-0008-2025:t{table_number}:"
                        f"r{row_number:03d}:{_slug(product_name)}:l{level}"
                    ),
                    product_name=product_name,
                    product_group=path[0] if len(path) > 1 else "",
                    head_material=head_material,
                    production_process=process,
                    product_form=identity.product_form or "",
                    quota_level=level,
                    value_kgce_per_t=value,
                    standard_code=STANDARD_CODE,
                    table_number=table_number,
                    physical_page=page_index + 1,
                    printed_page=page_index - 3,
                    canonical_product=canonical_product_key(product_name),
                    applicability=product_path,
                    source_locator=locator,
                    source_sha256=source_sha,
                ))
    return output


def _coefficient_values(value: str) -> tuple[float, float] | None:
    compact = re.sub(r"(?<=\d)\s+(?=\d)", "", clean_cell(value))
    observed = [float(item) for item in re.findall(r"(\d+(?:\.\d+)?)\s*kgce", compact)]
    if not observed:
        return None
    return min(observed), max(observed)


def _coefficient_unit(value: str) -> str:
    observed = clean_cell(value).casefold().replace(" ", "")
    if "kw" in observed:
        return "kgce/kWh"
    if "/m3" in observed or "/m³" in observed:
        return "kgce/m3"
    if "/kg" in observed:
        return "kgce/kg"
    if "/mj" in observed:
        return "kgce/MJ"
    if "/t" in observed:
        return "kgce/t"
    raise ValueError(f"unsupported standard-coal conversion unit: {value}")


def extract_conversions(pdf: pdfplumber.PDF) -> list[EnergyConversionRecord]:
    output: list[EnergyConversionRecord] = []
    for page_index, table_number in ((11, "A.1"), (12, "B.1")):
        for table in pdf.pages[page_index].extract_tables():
            for raw in table:
                if len(raw) < 3:
                    continue
                name_parts = [clean_cell(item) for item in raw[:-2] if clean_cell(item)]
                coefficient = clean_cell(raw[-1])
                values = _coefficient_values(coefficient)
                if not name_parts or values is None:
                    continue
                carrier = name_parts[-1]
                if carrier in {"能源名称", "品 种"}:
                    continue
                parameter_name = {
                    "电力（当量值）": "electricity_kgce_per_kwh",
                    "天然气": "natural_gas_kgce_per_nm3",
                }.get(carrier, f"standard_coal_conversion_{_slug(carrier)}")
                source_locator = f"standard:{STANDARD_CODE.replace(' ', '-')}#table-{table_number}"
                output.append(EnergyConversionRecord(
                    conversion_id=f"t-chnrisc-0008-2025:{table_number}:{_slug(carrier)}",
                    parameter_name=parameter_name,
                    energy_carrier=carrier,
                    value_min=values[0],
                    value_max=values[1],
                    unit=_coefficient_unit(coefficient),
                    basis="equivalent_value" if "当量值" in carrier else "reference_standard_coal_coefficient",
                    source_type=ParameterSourceType.FORMAL_STANDARD,
                    provider=PUBLISHER,
                    locator=source_locator,
                    citation=f"{STANDARD_CODE} 表{table_number}，{carrier}",
                    quality_note=(
                        "exact published coefficient" if values[0] == values[1]
                        else "published range; runtime derivation requires an independently selected exact value"
                    ),
                    standard_code=STANDARD_CODE,
                    physical_page=page_index + 1,
                    metadata={"raw_coefficient": coefficient},
                ))
    return output


def extract_modifier_rules(pdf: pdfplumber.PDF) -> list[EnergyQuotaModifierRule]:
    output: list[EnergyQuotaModifierRule] = []
    sources = ((6, 0, "1"), (8, 0, "2"), (8, 1, "3"))
    for page_index, table_index, table_number in sources:
        tables = pdf.pages[page_index].extract_tables()
        note_cells = [
            str(cell)
            for row in tables[table_index]
            for cell in row
            if cell and "注：" in str(cell)
        ]
        note_block = note_cells[0].split("注：", 1)[1] if note_cells else ""
        parsed: list[tuple[str, str]] = []
        current_id = ""
        current_lines: list[str] = []
        last_number = 0
        for line in note_block.splitlines():
            observed = line.strip()
            match = re.match(r"^(\d+)\s*(.*)$", observed)
            number = int(match.group(1)) if match else 0
            description = match.group(2).strip() if match else ""
            starts_note = bool(
                match and 1 <= number <= 9 and number == last_number + 1
                and any(not char.isdigit() for char in description)
            )
            if starts_note:
                if current_id:
                    parsed.append((current_id, " ".join(current_lines)))
                current_id = str(number)
                current_lines = [description]
                last_number = number
            elif current_id and observed:
                current_lines.append(observed)
        if current_id:
            parsed.append((current_id, " ".join(current_lines)))
        for note_id, description in parsed:
            cleaned = " ".join(description.split())
            if table_number == "3" and note_id == "4":
                cleaned = cleaned.removesuffix(" 2 3").replace("Al0 含量", "Al2O3 含量")
            if not cleaned:
                continue
            output.append(EnergyQuotaModifierRule(
                rule_id=f"t-chnrisc-0008-2025:t{table_number}:note-{note_id}",
                standard_code=STANDARD_CODE,
                table_number=table_number,
                note_id=note_id,
                description=cleaned,
                adjustment_type="conditional_text",
                applicability="requires explicit user/process evidence before application",
                physical_page=page_index + 1,
            ))
    return output


def load_process_parameters(path: Path | None) -> list[ScopedProcessParameterRecord]:
    if path is None:
        return []
    values = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(values, list):
        raise ValueError("process-parameter JSON must be a list")
    return [ScopedProcessParameterRecord(
        parameter_id=item["parameter_id"],
        name=item["name"],
        value=float(item["value"]),
        unit=item["unit"],
        source_type=ParameterSourceType(item["source_type"]),
        provider=item["provider"],
        locator=item["locator"],
        reference_head_material=item["reference_head_material"],
        reference_process=item["reference_process"],
        target_head_material=item["target_head_material"],
        target_process=item["target_process"],
        reference_source_id=item.get("reference_source_id", ""),
        citation=item.get("citation", ""),
        quality_note=item.get("quality_note", ""),
        metadata=item.get("metadata", {}),
    ) for item in values]


def annotate_profile_quota_comparisons(
    profiles: list[EnterpriseEnergyProfileRecord],
    quotas: list[EnergyQuotaRecord],
) -> list[EnterpriseEnergyProfileRecord]:
    quota_groups: dict[tuple[str, int], list[EnergyQuotaRecord]] = {}
    for quota in quotas:
        key = (
            _workbook_text(quota.canonical_product or quota.product_name).casefold(),
            quota.quota_level,
        )
        quota_groups.setdefault(key, []).append(quota)
    output: list[EnterpriseEnergyProfileRecord] = []
    for profile in profiles:
        key = (_workbook_text(profile.canonical_product).casefold(), profile.quota_level)
        matches = quota_groups.get(key, [])
        metadata = dict(profile.metadata)
        if len(matches) == 1:
            quota = matches[0]
            difference = profile.total_energy_kgce_per_t - quota.value_kgce_per_t
            metadata.update({
                "formal_quota_record_id": quota.record_id,
                "formal_quota_value_kgce_per_t": f"{quota.value_kgce_per_t:g}",
                "workbook_minus_formal_quota_kgce_per_t": f"{difference:g}",
            })
            quality_note = profile.quality_note
            if abs(difference) > 1e-9:
                quality_note += (
                    f" | workbook energy differs from formal quota {quota.value_kgce_per_t:g} "
                    f"kgce/t by {difference:g} kgce/t"
                )
            output.append(replace(profile, metadata=metadata, quality_note=quality_note))
        else:
            metadata["formal_quota_match"] = "none" if not matches else "ambiguous"
            output.append(replace(profile, metadata=metadata))
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("pdf", type=Path)
    parser.add_argument("database", type=Path)
    parser.add_argument("--expected-source-sha256", required=True)
    parser.add_argument("--dataset-version")
    parser.add_argument("--process-parameters-json", type=Path)
    parser.add_argument("--enterprise-energy-workbook", type=Path)
    parser.add_argument("--expected-workbook-sha256")
    parser.add_argument("--overwrite", action="store_true")
    args = parser.parse_args()
    observed_sha = sha256(args.pdf)
    if observed_sha != args.expected_source_sha256.casefold():
        raise SystemExit(
            f"source PDF SHA-256 mismatch: expected {args.expected_source_sha256}, observed {observed_sha}"
        )
    with pdfplumber.open(args.pdf) as pdf:
        text = "\n".join((page.extract_text() or "") for page in pdf.pages[:5])
        if "T/CHNRISC 0008" not in text or "2025" not in text:
            raise SystemExit("source PDF is not the expected T/CHNRISC 0008-2025 standard")
        quotas = extract_quotas(pdf, observed_sha)
        conversions = extract_conversions(pdf)
        modifier_rules = extract_modifier_rules(pdf)
    enterprise_profiles: list[EnterpriseEnergyProfileRecord] = []
    enterprise_process_emissions: list[EnterpriseProcessEmissionRecord] = []
    workbook_sha = ""
    if args.enterprise_energy_workbook:
        if not args.expected_workbook_sha256:
            raise SystemExit(
                "--expected-workbook-sha256 is required with --enterprise-energy-workbook"
            )
        workbook_sha = sha256(args.enterprise_energy_workbook)
        if workbook_sha != args.expected_workbook_sha256.casefold():
            raise SystemExit(
                "source workbook SHA-256 mismatch: "
                f"expected {args.expected_workbook_sha256}, observed {workbook_sha}"
            )
        enterprise_profiles = extract_enterprise_profiles(
            args.enterprise_energy_workbook, workbook_sha
        )
        enterprise_profiles = annotate_profile_quota_comparisons(
            enterprise_profiles, quotas
        )
        enterprise_process_emissions = extract_enterprise_process_emissions(
            args.enterprise_energy_workbook, workbook_sha, enterprise_profiles
        )
    dataset_version = args.dataset_version or (
        "t-chnrisc-0008-2025+enterprise-energy-89/v3"
        if enterprise_profiles else "t-chnrisc-0008-2025/v1"
    )
    anchor = create_energy_database(
        args.database,
        database_name="refractory-energy-parameters.db",
        dataset_version=dataset_version,
        source_standard_code=STANDARD_CODE,
        source_sha256=observed_sha,
        source_locator=f"standard:{STANDARD_CODE.replace(' ', '-')}",
        quotas=quotas,
        conversions=conversions,
        process_parameters=load_process_parameters(args.process_parameters_json),
        enterprise_profiles=enterprise_profiles,
        enterprise_process_emissions=enterprise_process_emissions,
        modifier_rules=modifier_rules,
        additional_metadata=(
            {
                "enterprise_workbook_name": args.enterprise_energy_workbook.name,
                "enterprise_workbook_sha256": workbook_sha,
                "enterprise_workbook_sheet": ENTERPRISE_SHEET_NAME,
                "enterprise_workbook_range": "A1:W274",
                "enterprise_profile_records": str(len(enterprise_profiles)),
                "enterprise_process_emission_records": str(
                    len(enterprise_process_emissions)
                ),
                "energy_selection_policy_id": (
                    "process.database-priority-energy-replacement/v1"
                ),
            }
            if enterprise_profiles else {}
        ),
        overwrite=args.overwrite,
    )
    print(json.dumps({
        "anchor": anchor.to_dict(),
        "quota_records": len(quotas),
        "conversion_records": len(conversions),
        "modifier_rules": len(modifier_rules),
        "enterprise_profile_records": len(enterprise_profiles),
        "enterprise_process_emission_records": len(enterprise_process_emissions),
        "enterprise_runtime_eligible_records": sum(
            item.runtime_eligible for item in enterprise_profiles
        ),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    sys.exit(main())
